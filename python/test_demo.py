from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import  WebDriverWait #bekleme islemlerini yapar
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path #klasor olusturmak icin
from datetime import date
import openpyxl #exel icin
from constants import globalConstants
#prefix-->on ek e gore yani (test_)ile baslamali test fonk olarak gorur

class Test_DemoClass:
    def setup_method(self): #her testten once cagirilir
        self.driver=webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
#gunun tarihini al bu tarih ile bir klasor var mı kontrol et yoksa 
        self.folderPath=str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True) #exist_ok=True ilgili klasor varsa olusturma demek
    
    
    def teardown_method(self): #her testten sonra cagrilir
        self.driver.quit() 
    #setup->test_demoFunc->teardown
    def test_demoFunc (self):
        # 3A Act Arrange Assert
        text="Hello"
        assert text=="Hello"
    #setup->test_demo2->teardown
    def test_demo2(self):
        assert True
    
    def getData():  #decoratorlerden cagrilan fonksiyonlara self parametresi tanımlanmaz
       #veriyi al
        exelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet=exelFile["Sayfa1"] #dosyadaki hangi sayfayı okdugunu verir
        total_rows=selectedSheet.max_row
        data=[]
                        #toplam satır sayısı
        for i in range(2,total_rows+1):
            username=selectedSheet.cell(i,1).value
            password=selectedSheet.cell(i,2).value
            tupleData=(username,password)
            data.append(tupleData)

        return data

    #bu alttaki yorum satiri decoratordur
    #@pytest.mark.skip() bu fonk test ederken atla dedik

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput=self.driver.find_element(By.ID,"username")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput=self.driver.find_element(By.ID,"password")
        sleep(2)
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        sleep(2)
        loginBtn=self.driver.find_element(By.ID,"login-button")
        sleep(2)
        loginBtn.click()
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3/text()")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        #magic string
        assert errorMessage.text=="Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator)) #max 5 sn olcak sekilde ilgili locatorun gorunmesini bekle
